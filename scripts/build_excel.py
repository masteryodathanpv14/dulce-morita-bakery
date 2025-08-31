import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import datetime

BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "costeo-pasteleria"
OUT_XLSX = DATA_DIR / "modelo-costeo.xlsx"

def write_df(ws, df):
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=j, value=col)
    for i, record in enumerate(df.itertuples(index=False), start=2):
        for j, value in enumerate(record, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if isinstance(value, (datetime.date, datetime.datetime)):
                cell.number_format = "yyyy-mm-dd"

def create_table(ws, table_name):
    max_row = ws.max_row
    max_col = ws.max_column
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

def add_formula_column(ws, header, formula_by_row, start_row=2):
    col_idx = ws.max_column + 1
    ws.cell(row=1, column=col_idx, value=header)
    for r in range(start_row, ws.max_row + 1):
        ws.cell(row=r, column=col_idx, value="=" + formula_by_row(r))
    return col_idx

def build():
    # Load CSVs
    tInsumos = pd.read_csv(DATA_DIR / "Insumos.csv")
    tProductos = pd.read_csv(DATA_DIR / "Productos.csv")
    tReceta = pd.read_csv(DATA_DIR / "RecetaDetalle.csv")

    tCompras_path = DATA_DIR / "tCompras.csv"
    if tCompras_path.exists():
        tCompras = pd.read_csv(tCompras_path)
        if "Fecha" in tCompras.columns:
            tCompras["Fecha"] = pd.to_datetime(tCompras["Fecha"]).dt.date
    else:
        tCompras = pd.DataFrame(columns=["Codigo","Fecha","Cantidad_compra","Unidad_compra","Precio_compra_COP","Proveedor"])

    wb = Workbook()
    wb.remove(wb.active)

    # tCompras first (so references exist)
    ws_com = wb.create_sheet("tCompras")
    write_df(ws_com, tCompras)
    create_table(ws_com, "tCompras")

    # tInsumos
    ws_ins = wb.create_sheet("tInsumos")
    write_df(ws_ins, tInsumos)
    # Calculated columns inside the table
    add_formula_column(ws_ins, "Unidad_base", lambda r:
        'IF(OR([@Unidad_compra]="kg",[@Unidad_compra]="g"),"g",IF(OR([@Unidad_compra]="l",[@Unidad_compra]="ml"),"ml","u"))')
    add_formula_column(ws_ins, "Factor_a_base", lambda r:
        'IF([@Unidad_compra]="kg",1000,IF([@Unidad_compra]="l",1000,1))')
    add_formula_column(ws_ins, "Cant_base_compra", lambda r:
        '[@Cantidad_compra]*[@Factor_a_base]')
    add_formula_column(ws_ins, "Costo_unit_base", lambda r:
        'IF([@Cant_base_compra]>0,[@Precio_compra_COP]/[@Cant_base_compra],0)')
    add_formula_column(ws_ins, "Costo_utilizable_base", lambda r:
        'IF(1-([@Merma_%]/100)>0,[@Costo_unit_base]/(1-([@Merma_%]/100)),0)')
    add_formula_column(ws_ins, "Precio_vigente_COP", lambda r:
        'IFERROR(XLOOKUP(MAXIFS(tCompras[Fecha],tCompras[Codigo],[@Codigo]),FILTER(tCompras[Fecha],tCompras[Codigo]=[@Codigo]),FILTER(tCompras[Precio_compra_COP],tCompras[Codigo]=[@Codigo]),[@Precio_compra_COP]),[@Precio_compra_COP])')
    add_formula_column(ws_ins, "Costo_unit_base_vigente", lambda r:
        'IF([@Cant_base_compra]>0,[@Precio_vigente_COP]/[@Cant_base_compra],0)')
    add_formula_column(ws_ins, "Costo_utilizable_base_vigente", lambda r:
        'IF(1-([@Merma_%]/100)>0,[@Costo_unit_base_vigente]/(1-([@Merma_%]/100)),0)')
    create_table(ws_ins, "tInsumos")

    # tProductos
    ws_pro = wb.create_sheet("tProductos")
    write_df(ws_pro, tProductos)
    create_table(ws_pro, "tProductos")

    # tReceta
    ws_rec = wb.create_sheet("tReceta")
    write_df(ws_rec, tReceta)
    add_formula_column(ws_rec, "Unidad_base_insumo", lambda r:
        'XLOOKUP([@Codigo_Insumo],tInsumos[Codigo],tInsumos[Unidad_base])')
    add_formula_column(ws_rec, "Costo_unit_insumo_base", lambda r:
        'XLOOKUP([@Codigo_Insumo],tInsumos[Codigo],tInsumos[Costo_utilizable_base_vigente])')
    add_formula_column(ws_rec, "Factor_a_base", lambda r:
        'IF(AND([@Unidad_receta]="kg",[@Unidad_base_insumo]="g"),1000,IF(AND([@Unidad_receta]="l",[@Unidad_base_insumo]="ml"),1000,1))')
    add_formula_column(ws_rec, "Cant_base", lambda r:
        '[@Cantidad_receta]*[@Factor_a_base]')
    add_formula_column(ws_rec, "Costo_parcial", lambda r:
        '[@Cant_base]*[@Costo_unit_insumo_base]')
    create_table(ws_rec, "tReceta")

    # Costeo
    ws_cos = wb.create_sheet("Costeo")
    headers = [
        "Producto","Costo_materiales_lote","Rendimiento_unid","MO_lote",
        "Ind_lote","Empaque_unit_COP","Costo_total_unit","Margen_objetivo",
        "IVA_venta","Precio_sin_IVA","Precio_con_IVA","Precio_redondeado_100"
    ]
    for j, h in enumerate(headers, start=1):
        ws_cos.cell(row=1, column=j, value=h)

    productos = tProductos["Producto"].tolist()
    for i, prod in enumerate(productos, start=2):
        ws_cos.cell(row=i, column=1, value=prod)
        ws_cos.cell(row=i, column=2, value="=SUMIFS(tReceta[Costo_parcial],tReceta[Producto],[@Producto])")
        ws_cos.cell(row=i, column=3, value="=XLOOKUP([@Producto],tProductos[Producto],tProductos[Rendimiento_lote_unid])")
        ws_cos.cell(row=i, column=4, value="=XLOOKUP([@Producto],tProductos[Producto],tProductos[Tiempo_lote_min])/60*XLOOKUP([@Producto],tProductos[Producto],tProductos[Mano_obra_hora_COP])")
        ws_cos.cell(row=i, column=5, value="=XLOOKUP([@Producto],tProductos[Producto],tProductos[Tiempo_lote_min])/60*XLOOKUP([@Producto],tProductos[Producto],tProductos[Overhead_hora_COP])")
        ws_cos.cell(row=i, column=6, value="=XLOOKUP([@Producto],tProductos[Producto],tProductos[Empaque_unit_COP])")
        ws_cos.cell(row=i, column=7, value="=([@Costo_materiales_lote]+[@MO_lote]+[@Ind_lote]) / [@Rendimiento_unid] + [@Empaque_unit_COP]")
        ws_cos.cell(row=i, column=8, value="=XLOOKUP([@Producto],tProductos[Producto],tProductos[Margen_objetivo])")
        ws_cos.cell(row=i, column=9, value="=XLOOKUP([@Producto],tProductos[Producto],tProductos[IVA_venta])")
        ws_cos.cell(row=i, column=10, value="=[@Costo_total_unit]/(1-[@Margen_objetivo])")
        ws_cos.cell(row=i, column=11, value="=[@Precio_sin_IVA]*(1+[@IVA_venta])")
        ws_cos.cell(row=i, column=12, value="=ROUNDUP([@Precio_con_IVA],-2)")

    last_row = 1 + len(productos)
    ref = f"A1:{get_column_letter(len(headers))}{last_row}"
    table = Table(displayName="tCosteo", ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium6",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws_cos.add_table(table)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Workbook generated at {OUT_XLSX}")

if __name__ == "__main__":
    build()